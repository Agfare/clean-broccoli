from __future__ import annotations

from pathlib import Path
from typing import Generator, List, Optional, Tuple

import openpyxl

from app.services.parsers.base import ParseResult, Segment

KNOWN_LANG_CODES = {
    "en", "de", "fr", "es", "it", "pt", "nl", "ru", "pl", "cs", "sk", "hu",
    "ro", "bg", "hr", "sr", "uk", "tr", "ar", "zh", "ja", "ko", "th", "vi",
    "id", "ms", "hi", "fa", "he", "el", "sv", "da", "no", "fi", "et", "lv", "lt",
}


def detect_xls_languages(path: Path) -> list[str]:
    """Return sorted list of language codes detected from column headers."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return []
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first_row:
            return []
        langs: set[str] = set()
        for cell in first_row:
            if cell is None:
                continue
            val = str(cell).strip().lower()
            # Direct match: header IS a lang code
            primary = val.split("-")[0].split("_")[0]
            if primary in KNOWN_LANG_CODES:
                langs.add(primary)
                continue
            # Header contains a lang code as a word token (e.g. "source_en", "text de")
            for sep in ("_", "-", " ", "."):
                for part in val.split(sep):
                    if part in KNOWN_LANG_CODES:
                        langs.add(part)
                        break
        return sorted(langs)
    except Exception:
        return []


def _detect_columns(headers: List[Optional[str]], source_lang: str, target_lang: str) -> Tuple[int, int]:
    """
    Detect source and target column indices from header row.
    Returns (source_col_idx, target_col_idx) (0-based).
    Falls back to (0, 1) if no match found.
    """
    src_lang_lower = source_lang.lower().split("-")[0]
    tgt_lang_lower = target_lang.lower().split("-")[0]

    source_idx: Optional[int] = None
    target_idx: Optional[int] = None

    for i, h in enumerate(headers):
        if h is None:
            continue
        hl = h.lower()
        if source_idx is None and ("source" in hl or src_lang_lower in hl):
            source_idx = i
        if target_idx is None and ("target" in hl or tgt_lang_lower in hl):
            target_idx = i

    if source_idx is None or target_idx is None:
        return 0, 1

    return source_idx, target_idx


def _row_to_segment(
    row, src_col: int, tgt_col: int, row_num: int,
    source_lang: str, target_lang: str,
) -> Optional[Segment]:
    """Convert a raw openpyxl row to a Segment, or None if the row is blank."""
    def _cell(col_idx: int) -> str:
        if col_idx < len(row) and row[col_idx] is not None:
            return str(row[col_idx]).strip()
        return ""

    src = _cell(src_col)
    tgt = _cell(tgt_col)
    if not src and not tgt:
        return None
    return Segment(
        id=str(row_num),
        source=src,
        target=tgt,
        source_lang=source_lang,
        target_lang=target_lang,
    )


def iter_xls(
    path: Path,
    source_lang: str,
    target_lang: str,
    warnings: Optional[List[str]] = None,
    progress_callback=None,
) -> Generator[Segment, None, None]:
    """Yield Segment objects one at a time from an XLS/XLSX file.

    Uses openpyxl read-only mode so rows are streamed from disk — memory stays
    flat regardless of file size.  Appends parse warnings to *warnings*.
    """
    if warnings is None:
        warnings = []

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as e:
        warnings.append(f"Failed to open XLS file: {e}")
        return

    ws = wb.active
    if ws is None:
        warnings.append("No active sheet found")
        wb.close()
        return

    row_iter = ws.iter_rows(values_only=True)
    first_raw = next(row_iter, None)
    if first_raw is None:
        warnings.append("Sheet is empty")
        wb.close()
        return

    first_strs = [str(c) if c is not None else None for c in first_raw]
    has_header = any(
        c and isinstance(c, str) and not c.strip().lstrip("-").replace(".", "").isdigit()
        for c in first_strs
    )

    if has_header:
        src_col, tgt_col = _detect_columns(first_strs, source_lang, target_lang)
        start_row = 2
    else:
        src_col, tgt_col = 0, 1
        start_row = 1
        seg = _row_to_segment(first_raw, src_col, tgt_col, 1, source_lang, target_lang)
        if seg:
            yield seg

    count = 0
    for row_offset, row in enumerate(row_iter):
        row_num = start_row + row_offset
        seg = _row_to_segment(row, src_col, tgt_col, row_num, source_lang, target_lang)
        if seg:
            count += 1
            if progress_callback is not None and count % 5_000 == 0:
                progress_callback(count)
            yield seg

    wb.close()


def parse_xls(path: Path, source_lang: str, target_lang: str, progress_callback=None) -> ParseResult:
    """Parse XLS/XLSX and return all segments as a list.

    For large files prefer *iter_xls* which yields one Segment at a time.
    """
    warnings: List[str] = []
    segments = list(
        iter_xls(path, source_lang, target_lang, warnings=warnings, progress_callback=progress_callback)
    )
    return ParseResult(
        segments=segments,
        warnings=warnings,
        encoding_ok=True,
        source_lang=source_lang,
        target_lang=target_lang,
    )
