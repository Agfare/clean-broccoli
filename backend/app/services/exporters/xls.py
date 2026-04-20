from __future__ import annotations

from pathlib import Path
from typing import List

import openpyxl
from openpyxl.cell.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.parsers.base import QAIssue, Segment

# ---------------------------------------------------------------------------
# Fill / font constants
# ---------------------------------------------------------------------------

_RED_FILL    = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
_GREEN_FILL  = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_LEGEND_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

_HEADER_FONT = Font(color="FFFFFF", bold=True)
_BOLD_FONT   = Font(bold=True)

# Shared legend data — used by both CleanXlsWriter and QaXlsWriter
_LEGEND_COLORS = [
    (_RED_FILL,    "Red",    "Segment has one or more ERROR-level QA issues"),
    (_YELLOW_FILL, "Yellow", "Segment has one or more WARNING-level QA issues"),
    (_GREEN_FILL,  "Green",  "Segment passed all QA checks"),
]

_LEGEND_CHECKS = [
    ("tags",         "Inline tag count mismatch between source and target"),
    ("variables",    "Variable/placeholder mismatch"),
    ("numbers",      "Numeric value mismatch"),
    ("scripts",      "Unexpected character script for target language"),
    ("untranslated", "Target is empty or identical to source"),
    ("duplicate",    "Duplicate segment found"),
    ("mt_quality",   "MT quality score below threshold"),
]


# ---------------------------------------------------------------------------
# Legend helper (write-only workbooks)
# ---------------------------------------------------------------------------

def _add_legend_sheet(wb: openpyxl.Workbook) -> None:
    """Append a colour-legend sheet to *wb* (write-only mode)."""
    ls = wb.create_sheet(title="Legend")

    hc = WriteOnlyCell(ls, value="Color Legend")
    hc.font = _BOLD_FONT
    hc.fill = _LEGEND_HEADER_FILL
    ls.append([hc])

    for fill, name, desc in _LEGEND_COLORS:
        nc = WriteOnlyCell(ls, value=name)
        nc.fill = fill
        ls.append([nc, desc])

    ls.append([])

    tc = WriteOnlyCell(ls, value="QA Check Types")
    tc.font = _BOLD_FONT
    ls.append([tc])

    for check, desc in _LEGEND_CHECKS:
        ls.append([check, desc])


# ---------------------------------------------------------------------------
# Streaming context-manager writers
# ---------------------------------------------------------------------------

class CleanXlsWriter:
    """Context-manager streaming writer for clean-segments XLSX.

    Uses openpyxl write-only mode so each row is flushed to disk immediately —
    memory stays flat regardless of how many segments are written.

    Usage::

        with CleanXlsWriter(path) as w:
            for seg in segments:
                w.write(seg)
    """

    _HEADERS    = ["ID", "Source", "Target", "Source Lang", "Target Lang"]
    _COL_WIDTHS = [15, 50, 50, 14, 14]

    def __init__(self, path: Path) -> None:
        self._path = path
        self._wb = None
        self._ws = None

    def __enter__(self) -> "CleanXlsWriter":
        self._path.parent.mkdir(parents=True, exist_ok=True)
        self._wb = openpyxl.Workbook(write_only=True)
        self._ws = self._wb.create_sheet("Clean Segments")
        for i, w in enumerate(self._COL_WIDTHS, start=1):
            self._ws.column_dimensions[get_column_letter(i)].width = w
        header_cells = []
        for h in self._HEADERS:
            c = WriteOnlyCell(self._ws, value=h)
            c.fill = _HEADER_FILL
            c.font = _HEADER_FONT
            header_cells.append(c)
        self._ws.append(header_cells)
        return self

    def write(self, seg: Segment) -> None:
        self._ws.append([seg.id, seg.source, seg.target, seg.source_lang, seg.target_lang])

    def __exit__(self, *args) -> None:
        _add_legend_sheet(self._wb)
        self._wb.save(str(self._path))
        return False


class QaXlsWriter:
    """Context-manager streaming writer for QA-report XLSX.

    Each segment+issues pair is written row-by-row with colour-coding, so
    memory stays flat no matter how many segments there are.

    Usage::

        with QaXlsWriter(path) as w:
            for seg, issues in seg_issue_pairs:
                w.write(seg, issues)
    """

    _HEADERS    = ["ID", "Source", "Target", "Source Lang", "Target Lang",
                   "QA Issues", "Severity", "Issue Details"]
    _COL_WIDTHS = [15, 40, 40, 14, 14, 22, 12, 55]

    def __init__(self, path: Path) -> None:
        self._path = path
        self._wb = None
        self._ws = None

    def __enter__(self) -> "QaXlsWriter":
        self._path.parent.mkdir(parents=True, exist_ok=True)
        self._wb = openpyxl.Workbook(write_only=True)
        self._ws = self._wb.create_sheet("QA Report")
        for i, w in enumerate(self._COL_WIDTHS, start=1):
            self._ws.column_dimensions[get_column_letter(i)].width = w
        header_cells = []
        for h in self._HEADERS:
            c = WriteOnlyCell(self._ws, value=h)
            c.fill = _HEADER_FILL
            c.font = _HEADER_FONT
            header_cells.append(c)
        self._ws.append(header_cells)
        return self

    def write(self, seg: Segment, issues: List[QAIssue]) -> None:
        has_error   = any(i.severity == "error"   for i in issues)
        has_warning = any(i.severity == "warning" for i in issues)
        if has_error:
            row_fill, severity_label = _RED_FILL, "error"
        elif has_warning:
            row_fill, severity_label = _YELLOW_FILL, "warning"
        else:
            row_fill, severity_label = _GREEN_FILL, "clean"

        issue_types   = ", ".join(sorted({i.check   for i in issues})) if issues else ""
        issue_details = "; ".join(i.message          for i in issues)  if issues else ""

        values = [seg.id, seg.source, seg.target, seg.source_lang, seg.target_lang,
                  issue_types, severity_label, issue_details]
        row_cells = []
        for v in values:
            c = WriteOnlyCell(self._ws, value=v)
            c.fill = row_fill
            row_cells.append(c)
        self._ws.append(row_cells)

    def __exit__(self, *args) -> None:
        _add_legend_sheet(self._wb)
        self._wb.save(str(self._path))
        return False
